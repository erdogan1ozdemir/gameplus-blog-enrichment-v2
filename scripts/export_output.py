"""
Gameplus Blog Enrichment — Output Exporter
===========================================
Flexible exporter for enriched blog HTML. Default = Excel rollup (title → html),
but supports several formats so the skill user can ask for whatever they need.

USAGE (from a build script):
    from export_output import export
    export(items, fmt="excel", out_path="gameplus-blog-icerikler.xlsx")

`items` is a list of dicts:
    [{"title": "Remake Nedir? ...", "slug": "remake-nedir-...", "html": "<style>...</style><h1>..."}]

Formats:
    excel   (default) — one row per blog: Başlık | HTML (Part 1) | HTML (Part 2) | Slug | Karakter
    files             — one .html file per blog (body-only), named by slug
    files-preview     — one full standalone .html per blog (with <html><body> wrapper), openable in browser
    combined          — single .html file with all blogs stacked (for quick review)

Excel cell limit is 32,767 chars; long HTML is split across two columns.
"""
import re
import sys
import os

EXCEL_CELL_LIMIT = 32700  # safe margin under the hard 32767 limit


def minify_html(html):
    """Remove comments, collapse inter-tag whitespace, tighten inline style attrs.
    Keeps the <style> block intact (only collapses its whitespace lightly)."""
    # Protect <style>...</style> blocks from aggressive >\s+< collapsing
    styles = []
    def stash(m):
        styles.append(m.group(0))
        return f'@@STYLE{len(styles)-1}@@'
    html = re.sub(r'<style>.*?</style>', stash, html, flags=re.DOTALL)

    html = re.sub(r'<!--.*?-->', '', html, flags=re.DOTALL)
    html = re.sub(r'>\s+<', '><', html)

    def strip_style(m):
        s = m.group(1)
        s = re.sub(r'\s*:\s*', ':', s)
        s = re.sub(r'\s*;\s*', ';', s)
        s = re.sub(r',\s+', ',', s)
        return f'style="{s.strip()}"'
    html = re.sub(r'style="([^"]*)"', strip_style, html)
    html = re.sub(r'  +', ' ', html)

    # Restore style blocks (lightly minified)
    def restore(m):
        idx = int(m.group(1))
        block = styles[idx]
        block = re.sub(r'\n\s*', ' ', block)
        block = re.sub(r'\s{2,}', ' ', block)
        return block
    html = re.sub(r'@@STYLE(\d+)@@', restore, html)
    return html.strip()


def _split_for_cells(minified):
    """Split into a list of chunks each <= EXCEL_CELL_LIMIT, preferring tag boundaries.
    Concatenating the chunks in order reconstructs the original."""
    if len(minified) <= EXCEL_CELL_LIMIT:
        return [minified]
    chunks, s = [], minified
    while len(s) > EXCEL_CELL_LIMIT:
        window = s[EXCEL_CELL_LIMIT - 3000:EXCEL_CELL_LIMIT]
        ms = list(re.finditer(r'><', window))
        split = (EXCEL_CELL_LIMIT - 3000) + ms[-1].start() + 1 if ms else EXCEL_CELL_LIMIT
        chunks.append(s[:split]); s = s[split:]
    if s:
        chunks.append(s)
    return chunks


def export_excel(items, out_path):
    """One row per blog. HTML is minified then split across HTML 1..N columns
    (each <= 32.7k chars). Concatenate HTML 1..N in order to get the full body."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    rows, max_parts = [], 1
    for item in items:
        minified = minify_html(item['html'])
        parts = _split_for_cells(minified)
        max_parts = max(max_parts, len(parts))
        rows.append((item.get('title', ''), item.get('slug', ''), len(minified), parts))
    wb = Workbook()
    ws = wb.active
    ws.title = "Gameplus Blog İçerikleri"
    headers = ['Başlık', 'Slug', 'Karakter'] + ['HTML %d' % (i + 1) for i in range(max_parts)]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h).font = Font(bold=True, size=11)
    for r, (title, slug, n, parts) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=title)
        ws.cell(row=r, column=2, value=slug)
        ws.cell(row=r, column=3, value=n)
        for j, part in enumerate(parts):
            ws.cell(row=r, column=4 + j, value=part)
    ws.column_dimensions['A'].width = 46
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    for j in range(max_parts):
        ws.column_dimensions[chr(ord('D') + j)].width = 100
    ws.freeze_panes = 'A2'
    wb.save(out_path)
    return out_path


def export_files(items, out_dir, preview=False):
    os.makedirs(out_dir, exist_ok=True)
    written = []
    for item in items:
        slug = item.get('slug') or re.sub(r'[^a-z0-9]+', '-', item.get('title', 'blog').lower()).strip('-')[:60]
        html = item['html']
        suffix = '.html'
        if preview:
            # html is expected to already be a full standalone doc; if not, the build script wraps it
            fname = f"ornek-blog-{slug}{suffix}"
        else:
            fname = f"ornek-blog-{slug}-body-only{suffix}"
        path = os.path.join(out_dir, fname)
        with open(path, 'w', encoding='utf-8') as f:
            f.write(html)
        written.append(path)
    return written


def export_combined(items, out_path):
    parts = ['<!DOCTYPE html><html lang="tr"><head><meta charset="UTF-8">',
             '<meta name="viewport" content="width=device-width, initial-scale=1.0">',
             '<title>Gameplus Blog İçerikleri — Toplu Önizleme</title>',
             '<style>body{background:#000;color:#e5e7eb;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;max-width:900px;margin:0 auto;padding:24px 20px 60px;line-height:1.65;}',
             'a{color:#a3e635;} h1{color:#fff;} .gp-blog-sep{margin:60px 0;border:none;border-top:2px dashed #1f1f1f;}</style></head><body>']
    for i, item in enumerate(items):
        if i > 0:
            parts.append('<hr class="gp-blog-sep">')
        parts.append(item['html'])
    parts.append('</body></html>')
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(parts))
    return out_path


def export(items, fmt="excel", out_path=None, out_dir=None):
    """Main dispatcher. See module docstring for formats."""
    if fmt == "excel":
        out_path = out_path or "gameplus-blog-icerikler.xlsx"
        return export_excel(items, out_path)
    elif fmt == "files":
        out_dir = out_dir or "."
        return export_files(items, out_dir, preview=False)
    elif fmt == "files-preview":
        out_dir = out_dir or "."
        return export_files(items, out_dir, preview=True)
    elif fmt == "combined":
        out_path = out_path or "gameplus-blog-toplu-onizleme.html"
        return export_combined(items, out_path)
    else:
        raise ValueError(f"Unknown format: {fmt}. Use excel | files | files-preview | combined.")


if __name__ == "__main__":
    # Smoke test
    demo = [{"title": "Test Blog", "slug": "test-blog", "html": "<h1>Test</h1><p>İçerik</p>"}]
    print("minify:", minify_html(demo[0]["html"]))
    print("export excel →", export(demo, fmt="excel", out_path="/tmp/_smoke.xlsx"))
